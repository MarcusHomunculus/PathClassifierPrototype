from __future__ import annotations
from typing import Tuple, Iterator, Dict, Set, List
import re
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string

from matcher.enum.CellPropertyType import CellPropertyType
from matcher.xlsx.clustering.CellMatching import CellMatchingStruct, CellMatchResult
from matcher.xlsx.location.CellPositioning import CellPosition, CellPositionStruct, CellPositionStructType
from classifier.PathClassifier import PathClassifier
from classifier.error.MatchExceptions import ForwardFileNotFound
from matcher.clustering.ValueNamePair import ValueNamePair
from matcher.xlsx.clustering.CrossTableStruct import CrossTableStruct


class XlsxProcessor:

    FORWARDING_KEY = "forwarding_on"
    FORWARDING_PATH_KEY = "path_forward_symbol"
    WIDTH_USAGE_LIMITER = "width_only_in"
    TEMPLATE_CELL_ADDRESS_ROW_WISE = "${}{}:{}"
    TEMPLATE_CELL_ADDRESS_COL_WISE = "{}${}:{}"

    __classifier: PathClassifier
    __config = {}
    __root_xlsx: str
    __nested_xlsx_dir: str

    def __init__(self,
                 sink: PathClassifier,
                 config: Dict[str, str],
                 path_root_xlsx: str,
                 nested_xlsx_dir: str = "nested/"):
        """
        The constructor

        :param sink: the classifier to push the matches into
        :param config: a dictionary constructed from the config file
        :param path_root_xlsx: the path to the main Excel file
        :param nested_xlsx_dir: the path to the other Excel files which the root file might reference to
        """
        self.__classifier = sink
        self.__config = config
        self.__root_xlsx = path_root_xlsx
        root_file_name = re.search(r"\b\w*\.xlsx$", path_root_xlsx).group(0)
        root_path = path_root_xlsx[:-len(root_file_name)]
        if not nested_xlsx_dir.endswith("/"):
            nested_xlsx_dir += "/"
        self.__nested_xlsx_dir = root_path + nested_xlsx_dir

    def match_given_values_in(self, value_name_pairs: Iterator[ValueNamePair]) -> None:
        """
        Manages itself through the given xlsx-file and and tries to match the given pairs in the files (somewhere)

        :param value_name_pairs: a list of tuples with values and their corresponding URI
        """
        wb = load_workbook(self.__root_xlsx)
        for sheet in wb.sheetnames:
            self._check_row_wise(wb[sheet], value_name_pairs, self.__root_xlsx)
            self._check_column_wise(wb[sheet], value_name_pairs, self.__root_xlsx)
            self._check_as_cross_table(wb[sheet], value_name_pairs, self.__root_xlsx)

    def receive_for_path(self, path: str, name: str, nested_path: str = "") -> List[str]:
        """
        Resolves the given path and translates the name into an associated value (list) in the excel file

        :param path: the path to the information in the excel table
        :param name: the name to the value wanted
        :param nested_path: the path to use when following a file forwarding
        :return: the value addressed
        """
        path_parts = path.split(";")
        if len(path_parts) == 3:
            # is a cross-table -> no forwarding here so tackle the problem just straight on
            return self._get_from_cross_table(name, path_parts[0], path_parts[1], path_parts[2])
        if len(path_parts) == 2:
            # is a row- or column-table
            return self._get_from_linear_table(name, path_parts[0], path_parts[1], nested_path)
        else:
            raise AttributeError("Can't decode the type of the table the path '{}' represents".format(path))

    @staticmethod
    def extract_name_path(value_name_path: str) -> str:
        """
        Returns the name path of the given value-name path

        :param value_name_path: the path of a value-name combination
        :return: the "absolute" path to the name
        """
        path_parts = value_name_path.split(";")
        if len(path_parts) == 3:
            # means it is a cross matrix
            basic_path = XlsxProcessor.__extract_base_path(value_name_path)
            # only the first cell position has a '@' -> so prepend it
            name_position = path_parts[2] if path_parts[2].startswith("@") else ("@" + path_parts[2])
            return "{}/{}".format(basic_path, name_position)
        elif len(path_parts) == 2:
            # means it is a row- or column-table
            if not path_parts[1].startswith("@"):
                return path_parts[1]
            basic_path = XlsxProcessor.__extract_base_path(value_name_path)
            return "{}/{}".format(basic_path, path_parts[1])
        raise AttributeError("The given paths structure is unknown: {}. Giving up disassembling it".format(
            value_name_path))

    def get_names(self, sink_name_path: str) -> Set[str]:
        """
        Returns the list of names that can be taken from the given name path (**only**). If the given path does not
        address names the function might enter undefined behaviour

        :param sink_name_path: the path to the names
        :return: an set holding all names found
        """
        file_name, sheet_name = self.__disassemble_base_path(sink_name_path)
        wb = load_workbook(file_name)
        position, is_fixed_row = CellPosition.from_cell_path_position(sink_name_path)
        names = []
        # the name path will never have forwarding in this scenario so just interpret the next piece as sheet and check
        # the given positions content
        sheet = wb[sheet_name]
        for cell in XlsxProcessor.__get_cell_line_iterator(sheet, position, is_fixed_row):
            if cell.value is None:
                continue
            names.append(cell.value)
        return set(names)

    def _check_row_wise(self, sheet: Worksheet, value_name_pairs: Iterator[ValueNamePair],
                        path: str, check_for_value_only: bool = False) -> CellPositionStruct:
        """
        Iterates through the rows of the sheet given and tries to match the given list of value-URI-pairs in a row-wise
        fashion. If a match could be made the resulting path is pushed into the classifier

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        :param check_for_value_only: set this flag to return immediately after a value from the list is found
        :return a LineResultStruct if only a value has been found: this is important for forwarded searches else an
                invalid / empty struct
        """
        current_sheet_path = "{}/{}".format(path, sheet.title)
        lowest_header_row = -1
        forward_index = -1
        handle_forwarding, forwarding_column_name = self.__includes_forwarding(sheet.title)
        color_key = "header_{}".format(sheet.title)
        if color_key not in self.__config:
            raise AssertionError("Missing key '{}' for retrieving the header color".format(color_key))
        header_color = self.__config[color_key]
        row_index = 0   # start with zero to be conform with xlsx
        for row in sheet.iter_rows():
            row_index += 1
            result = self.__scan_cell_line_for(row, current_sheet_path, sheet, value_name_pairs, forward_index,
                                               header_color, "" if not handle_forwarding else forwarding_column_name)
            if result.read_result == CellPositionStructType.NO_FINDING:
                continue
            elif result.read_result == CellPositionStructType.HEADER_FOUND:
                lowest_header_row = row_index
                if result.contains_header_forwarding_position():
                    forward_index = column_index_from_string(result.name_or_forward_position.column)
                continue
            # else data has been found
            if result.match_struct.success_type == CellMatchResult.VALUE_FOUND and check_for_value_only:
                # only a value has been found -> forward the information to the caller -> but if no header has been
                # found the orientation is probably bogus
                if lowest_header_row < 0:
                    # probably wrong orientation
                    return CellPositionStruct.create_no_find()
                # add the final piece to the path
                result.value_path += self.__to_linear_cell_address(False, result.value_position.column,
                                                                   lowest_header_row + 1,
                                                                   result.value_position.read_type)
                return result
            elif lowest_header_row < 0:
                # means no header has been found
                continue
            elif result.match_struct.success_type != CellMatchResult.ALL_FOUND:
                # this case shouldn't exist in reality yet cover it to be on the safe side
                continue
            # else a pair has been detected -> collect all the required data and push the result into the classifier
            row_data_start = lowest_header_row + 1
            name_cell = self.__to_linear_cell_address(False, result.name_or_forward_position.column, row_data_start,
                                                      result.name_or_forward_position.read_type)
            # if the value path was external / forwarding use the one stored -> only the locals know the environment
            if result.contains_value_forwarding_path():
                final_path = result.value_path + ";"    # make the path final
                # prepend the current path to the name path to indicate that both differ
                name_cell = "{}/@{}".format(current_sheet_path, name_cell)
            else:
                value_cell = self.__to_linear_cell_address(False, result.value_position.column, row_data_start,
                                                           result.value_position.read_type)
                final_path = "{}/@{};".format(current_sheet_path, value_cell)
                name_cell = "@{}".format(name_cell)
            final_path += name_cell
            self.__classifier.add_potential_match(final_path)
        # if something was found it has been pushed to the classifier already -> so no need to return anything
        return CellPositionStruct.create_no_find()

    def _check_column_wise(self, sheet: Worksheet, value_name_pairs: Iterator[ValueNamePair],
                           path: str, check_for_value_only: bool = False) -> CellPositionStruct:
        """
        Iterates through the columns of the sheet given and tries to match the given list of value-URI-pairs in a
        column-wise fashion. If a match could be made the resulting path is pushed into the classifier

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        :param check_for_value_only: set this flag to return immediately after a value from the list is found
        :return a LineResultStruct if only a value has been found: this is important for forwarded searches else an
                invalid / empty struct
        """
        current_sheet_path = "{}/{}".format(path, sheet.title)
        lowest_header_col = -1
        forward_index = -1
        handle_forwarding, forwarding_row_name = self.__includes_forwarding(sheet.title)
        header_color = self.__config["header_{}".format(sheet.title)]
        col_index = -1
        for column in sheet.iter_cols():
            col_index += 1
            result = self.__scan_cell_line_for(column, current_sheet_path, sheet, value_name_pairs, forward_index,
                                               header_color, "" if not handle_forwarding else forwarding_row_name)
            if result.read_result == CellPositionStructType.NO_FINDING:
                continue
            elif result.read_result == CellPositionStructType.HEADER_FOUND:
                lowest_header_col = col_index
                if result.contains_header_forwarding_position():
                    forward_index = result.name_or_forward_position.row
                continue
            if result.match_struct.success_type == CellMatchResult.VALUE_FOUND and check_for_value_only:
                # only a value has been found -> forward the information to the caller -> but if no header has been
                # found the orientation is probably bogus
                if lowest_header_col < 0:
                    # probably wrong orientation
                    return CellPositionStruct.create_no_find()
                result.value_path += self.__to_linear_cell_address(True, get_column_letter(lowest_header_col + 1),
                                                                   result.value_position.row,
                                                                   result.value_position.read_type)
                return result
            elif lowest_header_col < 0:
                # means no header has been found
                continue
            elif result.match_struct.success_type != CellMatchResult.ALL_FOUND:
                # this case shouldn't exist in reality yet cover it to be on the safe side
                continue
            # else a pair has been detected -> collect all the required data and push the result into the classifier
            col_data_start = get_column_letter(lowest_header_col + 1)
            name_cell = self.__to_linear_cell_address(True, col_data_start, result.name_or_forward_position.row,
                                                      result.name_or_forward_position.read_type)
            if result.contains_value_forwarding_path():
                final_path = result.value_path + ";"
                # prepend the current path to the name path to indicate that both differ
                name_cell = "{}/@{}".format(current_sheet_path, name_cell)
            else:
                value_cell = self.__to_linear_cell_address(True, col_data_start, result.value_position.row,
                                                           result.value_position.read_type)
                final_path = "{}/@{};".format(current_sheet_path, value_cell)
                name_cell = "@{}".format(name_cell)
            final_path += name_cell
            self.__classifier.add_potential_match(final_path)
        # if something was found it has been pushed to the classifier already -> so no need to return anything
        return CellPositionStruct.create_no_find()

    def _check_as_cross_table(self, sheet: Worksheet, value_name_pairs: Iterator[ValueNamePair], path: str) -> None:
        """
        Iterates through the sheet and tries to determine if the sheet could contain a cross table where names and
        values are distributed along a column and a row and their clustering is indicated by a 'X' in the field below

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        """

        def find_xs(start_row: int, how_many: int) -> bool:
            """
            Goes through the rows below the given one and counts every single X it encounters: if the count reaches the
            how_many-argument true is returned

            :param start_row: the row to start searching at
            :param how_many: the many x's should be found
            :return: true if the expected count was reached else false
            """
            x_count = 0
            for row in sheet.iter_rows(min_row=start_row):
                for cell in row:
                    if cell.value == "x" or cell.value == "X":
                        x_count += 1
                    if x_count >= how_many:
                        return True
            return False

        def scan_columns_for_list(pair_list: Iterator[ValueNamePair]) -> Tuple[bool, CrossTableStruct]:
            """
            Checks the columns if either the list of values or names from the list of pairs can be found or if both
            lists are not present

            :param pair_list: the data points to look for
            :return: a tuple with true and the required data to continue or false and an useless struct if nothing could
                     be found
            """
            for col in sheet.iter_cols():
                list_found, first_data = CrossTableStruct.values_exist_in(col, pair_list)
                if list_found:
                    return True, first_data
            return False, CrossTableStruct()

        def scan_rows_for_list_until(data: CrossTableStruct) -> Tuple[bool, CrossTableStruct]:
            """
            Goes through the rows of the sheet in search for the list of the other part of the value-name pairs until
            the given row. Returns the given struct again to indicate that the struct might be altered

            :param data: the struct that holds all the required data
            :return: a tuple if enough values could be found and the updated struct in this case
            """
            for row in sheet.iter_rows(max_row=data.first_find.row):
                found_list = data.find_other_pair_values_in(row)
                if found_list:
                    # the second position had been stored with the check already
                    return True, data
            return False, data

        def find_data_field_start(to_scan: Iterator[Cell]) -> CellPosition:
            """
            Triggers on the first color change of cell background color that does come after a cell that was not white

            :param to_scan: the row / column to check for the color change
            :return: the position at which the transition from non-white to another color appeared
            """
            header_color = ""
            for cell in to_scan:
                if header_color == "" and self.__get_cell_color(cell) == "00000000":
                    continue    # ignore white cells which might be around the table itself
                elif self.__get_cell_color(cell) != header_color:
                    if header_color == "":
                        header_color = self.__get_cell_color(cell)
                    else:
                        # color has already been set so there's a new color at hand -> table header is reached
                        return CellPosition.create_from(cell)
            return CellPosition.create_invalid()

        # in a cross table everything should just stand in the table: so check if the values or the pairs can be found
        # in a column and go from there
        success, progress_struct = scan_columns_for_list(value_name_pairs)
        if not success:
            return
        # a cross matrix should have a side and a top -> now check if the top exists: it should be above the first value
        success, progress_struct = scan_rows_for_list_until(progress_struct)
        if not success:
            # probably not a cross table -> stop here
            return
        # now check if we find some "x" below which is a strong indicator in combination with the results above
        success = find_xs(progress_struct.opposite_find.row,
                          CrossTableStruct.REQUIRED_SUCCESS_RATE * progress_struct.get_sample_size())
        if not success:
            # no cross table after all
            return
        # if this line has been reached it is safe to assume to have a cross table at hand
        field_start_side = find_data_field_start(sheet[progress_struct.first_find.row])
        field_start_top = find_data_field_start(sheet[progress_struct.opposite_find.column])
        cross_area = "{}{}".format(field_start_side.column, field_start_top.row)
        top_area_template = "{}${}"
        side_area_template = "${}{}"
        if progress_struct.first_position_represents_value():
            value_area = side_area_template.format(progress_struct.first_find.column, field_start_top.row)
            name_area = top_area_template.format(field_start_side.column, progress_struct.opposite_find.row)
        else:
            value_area = top_area_template.format(field_start_side.column, progress_struct.opposite_find.row)
            name_area = side_area_template.format(progress_struct.first_find.column, field_start_top.row)
        final_path = "{}/{}/@{};{};{}".format(path, sheet.title, cross_area, value_area, name_area)
        self.__classifier.add_potential_match(final_path)

    def _follow_forward_to(self, file_name: str, work_path: str, testing_struct: CellMatchingStruct,
                           forwarding_index: int) -> CellPositionStruct:
        """
        Performs a row- and a column-wise search for the value in the file specified

        :param file_name: the file name to check for the value
        :param work_path: the internal path which describes the "address" from the root-file
        :param testing_struct: the struct which can be used to find the value
        :param forwarding_index: the index of the column or row that triggered the forwarding
        :return: A CellPositionStruct of "type" value_found or no_find
        """
        if testing_struct.success_type == CellMatchResult.NO_FINDING:
            # it makes no sense to look in another file without a previously found name
            return CellPositionStruct.create_no_find()
        file_path = self.__nested_xlsx_dir + file_name
        if not os.path.exists(file_path):
            raise ForwardFileNotFound("Could not find {} file under: {}".format(file_name, file_path))
        path = work_path + "/{}{}".format(self.__config[self.FORWARDING_PATH_KEY], forwarding_index)
        # create a dummy list which only contains the missing entry -> which has to be value else the forwarding would
        # be stupid
        value_pair: Iterator[ValueNamePair] = [ValueNamePair.create_with_value(testing_struct.get_missing_entry())]
        wb = load_workbook(file_path)
        for sheet in wb.sheetnames:
            # return the first value found
            result_row = self._check_row_wise(wb[sheet], value_pair, path, True)
            if result_row.contains_value_forwarding_path():
                return result_row
            result_col = self._check_column_wise(wb[sheet], value_pair, path, True)
            if result_col.contains_value_forwarding_path():
                return result_col
        return CellPositionStruct.create_no_find()

    def _get_from_cross_table(self, to_find: str, cross_area_path: str, value_path: str, name_path: str) -> List[str]:
        """
        Treats the sheet under the given name as cross table and extracts the data of it by checking for the name
        (to_find) and associate all values marked with "x" with it

        :param to_find: the name in question
        :param cross_area_path: the path that addresses the cross area. This path should also contain the file path
        :param value_path: the path to the first value cell which should only consist of the value cell position
        :param name_path: the path to the first name cell which should only consist of the name cell position
        :return: all values that could be associated with the given name
        """
        values_tuple = CellPosition.from_cell_path_position(value_path)
        names_tuple = CellPosition.from_cell_path_position(name_path)
        cross_area = CellPosition.from_cell_path_position(cross_area_path)[0]
        file_name, sheet_name = self.__disassemble_base_path(cross_area_path)
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        name_position, is_fixed_row = names_tuple
        for cell in XlsxProcessor.__get_cell_line_iterator(sheet, name_position, is_fixed_row):
            if cell.value != to_find:
                continue
            if is_fixed_row:
                cross_area.column = cell.column_letter
            else:
                cross_area.row = cell.row
            break

        value_template, name_is_fixed_row = values_tuple
        value_list = []
        for cross_cell in XlsxProcessor.__get_cell_line_iterator(sheet, cross_area, not is_fixed_row):
            if cross_cell.value is None:
                continue
            if "x" not in cross_cell.value and "X" not in cross_cell.value:
                continue
            # find the corresponding name
            if name_is_fixed_row:
                value_position = CellPosition(value_template.row, cross_cell.column, CellPropertyType.CONTENT)
            else:
                value_position = CellPosition(cross_cell.row, value_template.column, CellPropertyType.CONTENT)
            value = sheet[value_position.to_xlsx_position()].value
            value_list.append(value)
        return value_list

    def _get_from_linear_table(self, to_find: str, value_path: str, name_path: str, nested_path: str) -> List[str]:
        """
        Returns the value(s) from the path data provided for the name given

        :param to_find: the name to get the corresponding value of
        :param value_path: the path to the values in excel
        :param name_path: the path to the name in excel
        :param nested_path: the path to follow when accessing a file forwarding
        :return: a list of values found corresponding to the name (might also be a list of one element)
        """
        def contains_forwarding_at(to_check: List[str]) -> int:
            """
            Checks if the given path node lists contains a forwarding command and at which node index

            :param to_check: the list of path nodes to analyse
            :return: the index of the (first) node that contains a forwarding symbol
            """
            for i in range(len(to_check)):
                if self.__config[self.FORWARDING_PATH_KEY] in to_check[i]:
                    return i
            return -1

        def extract_value_same_sheet(work_sheet: Worksheet, value_start: CellPosition, name_start: CellPosition,
                                     fixed_row: bool) -> str:
            """
            Extracts the value to the given name from the table the sheet represents

            :param work_sheet: the sheet to check for the name (and therefor value)
            :param value_start: the cell which indicates from where the values start
            :param name_start: the position from where the names start
            :param fixed_row: if search has to be performed column wise or row-wise (in case of false)
            :return: a list containing one entry which matches the name
            """
            for cell in self.__get_cell_line_iterator(work_sheet, name_start, fixed_row):
                if cell.value is None or cell.value != to_find:
                    continue
                if fixed_row:
                    new_value_position = "{}{}".format(cell.column_letter, value_start.row)
                else:
                    new_value_position = "{}{}".format(value_start.column, cell.row)
                value_cell = work_sheet[new_value_position]
                if value_start.read_type == CellPropertyType.CONTENT and value_cell.value is not None:
                    return value_cell.value
                elif value_start.read_type == CellPropertyType.WIDTH:
                    return str(self.__get_cell_size(work_sheet, cell))
                else:
                    raise AttributeError("Can't decode type: {}. Are the if-branches out-dated?".format(
                        value_start.read_type))
            return ""

        def extract_value_list(work_sheet: Worksheet, value_start: CellPosition, fixed_row: bool):
            """
            Extracts all values that can be found in the given sheet from the given position on

            :param work_sheet: the table to read
            :param value_start: the position to start reading from
            :param fixed_row: if to read column wise or row wise
            :return: all values that exists under the value position
            """
            values: List[str] = []
            for cell in self.__get_cell_line_iterator(work_sheet, value_start, fixed_row):
                if cell.value is None:
                    continue
                if value_start.read_type == CellPropertyType.CONTENT:
                    values.append(cell.value)
                elif value_start.read_type == CellPropertyType.WIDTH:
                    values.append(str(self.__get_cell_size(work_sheet, cell)))
                else:
                    raise AttributeError("Can't decode type: {}. Are the if-branches out-dated?".format(
                        value_start.read_type))
            return values

        def extract_forward_index(forward_node: str) -> int:
            """
            Returns the index that is encoded in the forwarding identifier

            :param forward_node: the node holding the forward identifier
            :return: the index it holds
            """
            index_str = forward_node[len(self.__config[self.FORWARDING_PATH_KEY]):]
            return int(index_str)

        value_path_nodes = self.__disassemble_base_path(value_path)
        forwarding_node_index = contains_forwarding_at(value_path_nodes)
        if forwarding_node_index == -1:
            # means no forwarding is present -> all data can be found in one table
            wb = load_workbook(value_path_nodes[0])
            sheet = wb[value_path_nodes[1]]
            value_position, is_fixed_row = CellPosition.from_cell_path_position(value_path)
            name_position, _ = CellPosition.from_cell_path_position(name_path)
            result = extract_value_same_sheet(sheet, value_position, name_position, is_fixed_row)
            if not result:
                # this is a real error 'cause either the name list is inhomogeneous or the path is incorrect which means
                # the path training failed
                raise AttributeError("Could not extract a value from {};{}".format(value_path, name_path))
            # there can be only one value by this type of list -> wrap it in the list to comply to return value of other
            # functions
            return [result]
        else:
            # start with tracing the name and work from there
            name_path_nodes = self.__disassemble_base_path(name_path)
            name_position, is_fixed_row = CellPosition.from_cell_path_position(name_path)
            wb = load_workbook(name_path_nodes[0])
            sheet = wb[name_path_nodes[1]]
            forwarding_index = extract_forward_index(value_path_nodes[forwarding_node_index])
            if is_fixed_row:
                dummy_position = CellPosition(forwarding_index, name_position.column, CellPropertyType.CONTENT)
            else:
                dummy_position = CellPosition(name_position.row, get_column_letter(forwarding_index),
                                              CellPropertyType.CONTENT)
            # continue with extracting the values -> extract all at once -> this could also be done by returning an
            # iterator but this is overkill in this scenario
            file_name = extract_value_same_sheet(sheet, dummy_position, name_position, is_fixed_row)
            if nested_path:
                # prepend the required path if one is set
                file_name = nested_path + file_name
            wb = load_workbook(file_name)
            # the sheet name of the forwarding path comes after the forwarding symbol
            sheet = wb[value_path_nodes[forwarding_node_index + 1]]
            value_position, is_fixed_row = CellPosition.from_cell_path_position(value_path)
            return extract_value_list(sheet, value_position, is_fixed_row)

    def __includes_forwarding(self, sheet_name: str) -> Tuple[bool, str]:
        """
        Checks if the given sheet name is registered with a column which forwards to another file

        :param sheet_name: the name to check for
        :return: a tuple which contains if forwarding is to be expected and the column which is to be used for that
        """
        # as this project has only one root file the separation is simple
        names = self.__config[self.FORWARDING_KEY].split('/')
        if len(names) != 2:
            raise ValueError("The config file is wrong: expecting a path made of sheet/column. Received: " +
                             self.__config[self.FORWARDING_KEY])
        if names[0] != sheet_name:
            return False, ""
        return True, names[1]

    def __scan_cell_line_for(self, to_scan: Iterator[Cell], current_path: str, sheet: Worksheet,
                             value_name_pairs: Iterator[ValueNamePair] = None, forward_index: int = -1,
                             header_color: str = "", forward_name: str = "") -> CellPositionStruct:
        """
        Goes through the iteration of cells and checks if it can find useful information in it. The result is returned
        in form of struct which holds data depending on the found data

        :param to_scan: the iterator for a collection of cells (which is usually either a column or a row)
        :param current_path: the current path within the file / sheet structure
        :param sheet: the current worksheet to access the list of merged cells
        :param value_name_pairs: the values and names to find in the sheet. Set if you want to find these data
        :param forward_index: set this parameter if at this given index the content has to be interpreted as file name
        :param header_color: set this value if a header has to be detected
        :param forward_name: set this value if a header field containing this value indicates forwarding
        :return: a situation dependent initialized instance of LineResultStruct
        """
        def has_header_color(to_check: Cell) -> bool:
            """
            Checks if the given cell has a background color equal to the headers background color

            :param to_check: the cell to check for coloring
            :return: true if the colors match else false
            """
            return XlsxProcessor.__get_cell_color(to_check) == header_color

        if value_name_pairs is None:
            value_name_pairs = []
        current_idx = 0     # which results in starting the iteration with 1 which is the start for excel
        result_struct = CellMatchingStruct(value_name_pairs)
        value_position = CellPosition.create_invalid()
        name_position = CellPosition.create_invalid()
        value_path = ""
        is_header = False
        for cell in to_scan:
            current_idx += 1
            if cell.value is None:
                # skip the empty cell
                continue
            if has_header_color(cell):
                if cell.value == forward_name:
                    # return immediately as only one forwarding index is expected
                    # -> if multiple are required use a state machine
                    return CellPositionStruct.create_header_found(CellPosition.create_from(cell,
                                                                                           CellPropertyType.CONTENT))
                is_header = True
                continue
            # now the cell should contain some data -> find out if it is data of interest
            if current_idx == forward_index:
                result = self._follow_forward_to(cell.value, current_path, result_struct, current_idx)
                # write directly to value_position as the name shouldn't be found in the forwarding file
                # -> this would beat the whole purpose of the forwarding
                result_struct = result.match_struct
                # manually override the success-type as by copying it is only set the value found while no both are
                # available now
                result_struct.success_type = CellMatchResult.ALL_FOUND
                value_position = result.value_position
                value_path = result.value_path
            else:
                cell_data = self.__extract_cell_properties(cell, sheet)
                for data in cell_data.keys():
                    result = result_struct.test_value(data)
                    if result == CellMatchResult.NAME_FOUND:
                        name_position = CellPosition.create_from(cell, cell_data[data])
                    elif result == CellMatchResult.VALUE_FOUND:
                        value_position = CellPosition.create_from(cell, cell_data[data])
            if name_position.is_valid() and value_position.is_valid():
                # no reason to continue -> everything has been found
                return CellPositionStruct.create_data_pair_found(result_struct, value_position, name_position,
                                                                 value_path)
        # keep this separated as their separation ensures that header and data are detected in different lines
        # -> prefer the the header over the value as this ensures that header and value can only be detected in
        # different calls of the function and the header has to be detected first anyway
        if is_header:
            return CellPositionStruct.create_header_found(CellPosition.create_invalid())
        if value_position.is_valid():
            # assemble the value path -> this can't be completed however as higher level information is required
            value_path = "{}/@".format(current_path)
            return CellPositionStruct.create_value_found(result_struct, value_position, value_path)
        return CellPositionStruct.create_no_find()

    def __extract_cell_properties(self, to_extract_from: Cell, sheet: Worksheet) -> Dict[str, CellPropertyType]:
        """
        Takes the cell an creates a list of properties from it

        :param to_extract_from: the cell the properties are wanted from
        :return: a list of all properties supported
        """
        to_return = {to_extract_from.value: CellPropertyType.CONTENT}
        if self.WIDTH_USAGE_LIMITER not in self.__config or sheet.title in self.__config[self.WIDTH_USAGE_LIMITER]:
            # then add the width property
            to_return[str(self.__get_cell_size(sheet, to_extract_from))] = CellPropertyType.WIDTH
        return to_return

    @staticmethod
    def __get_cell_color(cell: Cell) -> str:
        """
        Returns the background color of the cell given

        :param cell: the cell in question
        :return: the ARGB of the cell color as string
        """
        return cell.fill.start_color.index

    @staticmethod
    def __to_linear_cell_address(is_fixed_row: bool, col: str, row: int, property_identifier: CellPropertyType) -> str:
        """
        Inserts the given arguments into the template selected depending on the first argument

        :param is_fixed_row: if the template should assume a row-wise reading scheme or a column-wise scheme
        :param col: the column to insert
        :param row: the row to insert
        :param property_identifier: the identifier of the cell property to use as source
        :return: a properly formatted string to be used as cell address
        """
        if is_fixed_row:
            template = XlsxProcessor.TEMPLATE_CELL_ADDRESS_COL_WISE
        else:
            template = XlsxProcessor.TEMPLATE_CELL_ADDRESS_ROW_WISE
        return template.format(col, row, str(property_identifier))

    @staticmethod
    def __extract_base_path(to_extract_from: str) -> str:
        """
        Extracts the path up to (but excluding) the cell position from the path given

        :param to_extract_from: the path to cut the position away from
        :return: the cell path except for the final part (the cell position)
        """
        return re.match(r"^.*(?=/@)", to_extract_from).group()

    @staticmethod
    def __disassemble_base_path(to_extract_from: str) -> List[str]:
        """
        Extracts the path up to (but excluding) the first cell position and splits it into their path nodes

        :param to_extract_from: the path to disassemble
        :return: a list of all path nodes except for the first cell position
        """
        # use substring to cut away the trailing '/'
        reduced_path = XlsxProcessor.__extract_base_path(to_extract_from)
        file_path = XlsxProcessor.__extract_file_path(reduced_path)
        to_return = [file_path]
        to_return.extend(reduced_path[len(file_path) + 1:].split("/"))
        return to_return

    @staticmethod
    def __extract_file_path(path: str) -> str:
        """
        Takes the given path and extracts the file path from it

        :param path: a potentially fully qualified path up to a cell position
        :return: only the part of the path that addresses the xlsx-file
        """
        result = re.search(r"^[\w./]*\.xlsx", path)
        if not result:
            raise AttributeError("Could not extract a file path from " + path)
        return result.group(0)

    @staticmethod
    def __get_cell_line_iterator(sheet: Worksheet, start: CellPosition, is_fixed_row: bool) -> Iterator[Cell]:
        """
        Returns an iterator for the cells starting from the given position in the sheet given.

        :param sheet: the sheet the cell line is wanted for
        :param start: the starting cell for the iteration
        :param is_fixed_row: if the line shall expand to a row or a column
        :return: the cells in the given row or column
        """
        if is_fixed_row:
            line = sheet.iter_cols(min_col=column_index_from_string(start.column), min_row=start.row, max_row=start.row)
        else:
            line = sheet.iter_rows(min_row=start.row, min_col=column_index_from_string(start.column),
                                   max_col=column_index_from_string(start.column))
        return [x[0] for x in line]

    @staticmethod
    def __get_cell_size(parent: Worksheet, to_read_from: Cell) -> int:
        """
        Returns the size of the cell at hand in the count of columns

        :param parent: the sheet that holds the cell in question
        :param to_read_from: the cell to get the size from
        :return: the size of the cell if it is merged else 1 is returned
        """
        # courtesy goes to: https://stackoverflow.com/a/57525843
        cell = parent.cell(to_read_from.row, to_read_from.column)
        for merged_cells in parent.merged_cells.ranges:
            if cell.coordinate in merged_cells:
                # as merged cells only expand in columns
                return merged_cells.max_col - merged_cells.min_col + 1
        return 1
