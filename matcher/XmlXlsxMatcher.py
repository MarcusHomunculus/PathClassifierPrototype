from typing import List, Tuple, Iterator, Dict
from enum import Enum
import xml.etree.ElementTree as ElemTree
import toml
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from classifier.BinClassifier import BinClassifier


class XmlXlsxMatcher:

    class HitType(Enum):
        NO_HIT = 0
        NAME_HIT = 1
        VALUE_HIT = 2

    class CellMatchStruct:
        success: bool
        expected: str
        found_one_is_name: bool

        def __init__(self, success: bool, expected: str = "", found_name: bool = False):
            """
            The constructor
            :param success: if the search for a value or name yielded a hit
            :param expected: the value or name to be found to match it to the content found
            :param found_name: true if the content found was the name else the value has been found
            """
            # start with a sanity check
            if success and expected == "":
                raise ValueError("Received an empty expected value were a string must be")
            self.success = success
            self.expected = expected
            self.found_one_is_name = found_name

    FORWARDING_KEY = "forwarding_on"
    TEMPLATE_CELL_ADDRESS_ROW_WISE = "${}{}:{}"
    TEMPLATE_CELL_ADDRESS_COL_WISE = "{}${}:{}"
    CELL_PROPERTY_CONTENT = "c"
    CELL_PROPERTY_WIDTH = "w"

    __config = {}
    __classifier = BinClassifier()
    __root_xlsx: str
    __nested_xlsx_dir: str

    def __init__(self, path_to_config: str, path_root_xlsx: str, nested_xlsx_dir: str = "nested/"):
        """
        The constructor

        :param path_to_config: the path to the config file
        :param path_root_xlsx: the path to the main Excel file
        :param nested_xlsx_dir: the path to the other Excel files which the root file might reference to
        """
        self.__read_config(path_to_config)
        self.__root_xlsx = path_root_xlsx
        root_file_name = re.search(r"\b\w*\.xlsx$", path_root_xlsx).group(0)
        root_path = path_root_xlsx[:-len(root_file_name)]
        if not nested_xlsx_dir.endswith("/"):
            nested_xlsx_dir += "/"
        self.__nested_xlsx_dir = root_path + nested_xlsx_dir

    def train(self, path_to_master_xml: str, path_to_slave_xlsx) -> None:
        # TODO: doc me
        tree = ElemTree.parse(path_to_master_xml)
        root = tree.getroot()
        for node in self._get_main_nodes():
            list_root = root.findall(".//{}".format(node))
            self._process_xml_master_nodes(list_root)

    def _axis_is_forwarding(self, name: str) -> bool:
        """
        Returns if the column (or row) references another Excel file which might be followed
        :param name: the identifier of the axis (column or row)
        :return: true if the program should open the file and search for data there
        """
        if self.FORWARDING_KEY not in self.__config:
            return False    # no forwarding specified
        if self.__config[self.FORWARDING_KEY] != name:
            return False
        return True

    def _process_xml_master_nodes(self, parent_node: ElemTree.Element):
        """
        Goes through the child list of the given node and treats them as master: meaning that the classifier will treat
        the xlsx as slave where the values from the xml has to be found in

        :param parent_node: the node which contains the list of nodes to match with the xlsx
        :return:
        """
        def process_attributes(node: ElemTree.Element, current_path: str, ids: List[str]) -> None:
            """
            Forwards the given attributes and their name to the function which tries to come up with matches

            :param node: the node which hosts the attributes
            :param current_path: the path to current node
            :param ids: the list of URI of all devices
            """
            for key in node.attr.keys():
                new_path = current_path + "/@{}".format(key)
                self.__classifier.add_source_path(new_path)
                values = self._path_to_xml_values(new_path, parent_node)
                pairs = zip(values, ids)    # hope that it fails in case both lists are not equal in length
                self._match_in_xslx(pairs)

        def process_node(node: ElemTree.Element, current_path: str, ids: List[str]) -> None:
            """
            Checks the given node for a value and attributes and forwards them to the function matching them to their
            counterpart in the xlsx. If the node contains child-nodes it processes recursively

            :param node: the node to extract the data (and children) from
            :param current_path: the path to the current node
            :param ids: the list of URI of all devices
            """
            if not node.text.isspace():
                self.__classifier.add_source_path(current_path)
                values = self._path_to_xml_values(current_path, parent_node)
                pairs = zip(values, ids)
                self._match_in_xslx(pairs)
            if node.attrib:
                process_attributes(node, current_path, ids)
            for child_node in node:
                name = child_node.tag
                process_node(child_node, current_path + "/{}".format(name), ids)

        identifier_list = []
        # get an overview about the targets to find
        for child in parent_node:
            # start with getting the identifier
            current_id = child.findall(".//{}".format(self._get_universal_id())).tag
            identifier_list.append(current_id)
        # use the first node as blue-print
        process_node(parent_node[0], "{}/{}".format(parent_node.tag, parent_node[0].tag), identifier_list)

    @staticmethod
    def _path_to_xml_values(path: str, root_node: ElemTree.Element) -> List[str]:
        """
        Resolves the given path to the nodes (or their attribute) of interest and returns the list of their values as
        they appear

        :param path: the path to resolve
        :param root_node: the node which contains the list of nodes eg. the node which is equivalent to the root of path
        :return: all values of the nodes or attributes under the given path
        """
        # as the path model is similar to XPath: just it
        result = re.search(r"(?<=@)\w*$", path)
        # cut away the root node
        root_end = path.index("/")
        node_path = path[root_end:]
        values = []
        if result is None:
            # means a node has to be processed
            nodes = root_node.findall(".{}".format(node_path))
            for node in nodes:
                values.append(node.text)
        else:
            # means a attribute has to be processed
            attribute_name = result.group(0)
            node_path = node_path[:-len(attribute_name) + 2]     # -1 for the "@" and -1 for the "/" before it
            nodes = root_node.findall(".{}".format(node_path))
            for node in nodes:
                values.append(node.attrib[attribute_name])
        return values

    def _match_values_to_xlsx_paths(self, value_name_pairs: Iterator[Tuple[str, str]], file_path: str) -> None:
        """
        Manages itself through the given xlsx-file and and tries to match the given pairs in the files (somewhere)

        :param value_name_pairs: a list of tuples with values and their corresponding URI
        :param file_path: the path to the file to match the values in
        """
        wb = load_workbook(file_path, True)
        sheet_names = wb.sheetnames
        for sheet in sheet_names:
            self.__search_sheet_for_value(value_name_pairs, wb[sheet], file_path)

    def _get_main_nodes(self) -> List[str]:
        """
        Extracts the names of the anchor nodes in the XML from the config

        :return: an iterable of the nodes which should be trained
        """
        raw_list = self.__config["List_nodes"]
        return raw_list.split(",")

    def _get_universal_id(self):
        """
        Returns the identifier which is used to distinguish the main nodes from each other

        :return: the identifier which can be used for matching
        """
        return self.__config["uri"]

    def __search_sheet_for_value(self, value_name_pairs: Iterator[Tuple[str, str]], sheet: Worksheet, path: str):
        """
        Analyses the given sheet in the way that it tries to find the given value-URI-pairs in all constellations it
        knows by just applying all and throw the result at the classifier

        :param value_name_pairs: a list of tuples with values and their corresponding URI
        :param sheet: the sheet to go through
        :param path: the current path to the sheet (for the classifier)
        """
        # usually a table is build column wise: so go through the sheet row wise to have a hit on the value and the
        # name
        self.__check_row_wise(sheet, value_name_pairs, path)
        # on good luck try it column wise
        self.__check_column_wise(sheet, value_name_pairs, path)

    def _match_in_xslx(self, values: Iterator[Tuple[str, str]]) -> None:
        """
        Starts the xlsx-reading process by analysing the main xlsx-file and manage its way through from there

        :param values: a list of value-URI-pairs to find in the xlsx-files (somewhere)
        """
        self._match_values_to_xlsx_paths(values, self.__root_xlsx)
        pass

    def __read_config(self, path_to_file: str) -> None:
        """
        Reads in the config file

        :param path_to_file: the path to the TOML file to read
        """
        config = {}
        with open(path_to_file, "r", encoding="utf-8") as c:
            config.update(toml.load(c))
        self.__config = config

    def __check_row_wise(self, sheet: Worksheet, value_name_pairs: Iterator[Tuple[str, str]], path: str) -> None:
        """
        Iterates through the rows of the sheet given and tries to match the given list of value-URI-pairs in a row-wise
        fashion. If a match could be made the resulting path is pushed into the classifier

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        """
        def get_data_start_of(column: Iterator[Cell]) -> int:
            """
            Iterates through the background color of the cells of the column given and returns the first row the
            (in the config file) specified header color is not used anymore

            :param column: the column to check for the first row of data lines
            :return: the row which should contain the first batch of data. If nothing could be found -1
            """
            header_hook = False
            header_color = self.__config["header_{}".format(sheet.title)]
            for item in column:
                if item.style.bg_color == header_color:
                    header_hook = True
                elif header_hook and item.style.bg_color != header_color:
                    return item.row
            return -1

        for values in sheet.iter_rows():
            # keep the index to check that the value belongs to the name and vice versa
            col_id = ""
            col_val = ""
            expected = ""  # an empty string resolves to false if converted to boolean
            for val in values:
                if not expected:
                    result = self.__match_cell_properties_to(val, value_name_pairs)
                    if result.success:
                        expected = result.expected
                        if result.found_one_is_name:
                            col_id = val.column_letter
                        else:
                            col_val = val.column_letter
                    else:   # this merely for the reader and not required by the syntax
                        continue
                # in the team file the cell holds the name and its size determines the size property -> so check this
                # one against the expected value, too
                if expected:
                    props = self.__extract_cell_properties(val)
                    for prop in props.keys():
                        if prop == expected:
                            col_id = val.column_letter if not col_id else col_id
                            col_val = val.column_letter if not col_val else col_val
                            # find the first line after the header once -> as the header should at one level the result
                            # should be true for both columns: so just pick one
                            row_data_start = get_data_start_of(sheet.iter_rows(col_id))
                            if row_data_start == -1:
                                # something went wrong: better abort
                                break
                            # assemble the path for the classifier it can match later on
                            final_path = "{}/{}/@{};{}".format(path, sheet.title,
                                                               self.__to_cell_address(
                                                                   self.TEMPLATE_CELL_ADDRESS_COL_WISE, col_val,
                                                                   row_data_start, props[prop]),
                                                               self.__to_cell_address(
                                                                   self.TEMPLATE_CELL_ADDRESS_COL_WISE, col_id,
                                                                   row_data_start, self.CELL_PROPERTY_CONTENT))
                            self.__classifier.add_potential_match(final_path)
                            return
        return

    def __check_column_wise(self, sheet: Worksheet, value_name_pairs: Iterator[Tuple[str, str]], path: str) -> None:
        """
        Iterates through the columns of the sheet given and tries to match the given list of value-URI-pairs in a
        column-wise fashion. If a match could be made the resulting path is pushed into the classifier

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        """

        def get_data_start_of(row: Iterator[Cell]) -> str:
            """
            Iterates through the background color of the cells of the column given and returns the first column the
            (in the config file) specified header color is not used anymore

            :param row: the row to check for the first column of data columns
            :return: the column which should contain the first batch of data. If nothing could be found an empty string
            """
            header_hook = False
            header_color = self.__config["header_{}".format(sheet.title)]
            for item in row:
                if item.style.bg_color == header_color:
                    header_hook = True
                elif header_hook and item.style.bg_color != header_color:
                    return item.column_letter
            return ""

        for values in sheet.iter_cols():
            # keep the index to check that the value belongs to the name and vice versa
            row_id = -1
            row_val = -1
            expected = ""  # an empty string resolves to false if converted to boolean
            for val in values:
                if not expected:
                    result = self.__match_cell_properties_to(val, value_name_pairs)
                    if result.success:
                        expected = result.expected
                        if result.found_one_is_name:
                            row_id = val.row
                        else:
                            row_val = val.row
                    else:  # this merely for the reader and not required by the syntax
                        continue
                # in the team file the cell holds the name and its size determines the size property -> so check this
                # one against the expected value, too
                if expected:
                    props = self.__extract_cell_properties(val)
                    for prop in props.keys():
                        if prop == expected:
                            row_id = val.row if not row_id else row_id
                            row_val = val.row if not row_val else row_val
                            # find the first column after the header once -> as the header should at one level the
                            # result should be true for both rows: so just pick one
                            col_data_start = get_data_start_of(sheet.iter_cols(row_id))
                            if not col_data_start:
                                # something went wrong: better abort
                                break
                            # assemble the path for the classifier it can match later on
                            final_path = "{}/{}/@{};{}".format(path, sheet.title,
                                                               self.__to_cell_address(
                                                                   self.TEMPLATE_CELL_ADDRESS_ROW_WISE, col_data_start,
                                                                   row_val, props[prop]),
                                                               self.__to_cell_address(
                                                                   self.TEMPLATE_CELL_ADDRESS_ROW_WISE, col_data_start,
                                                                   row_id, self.CELL_PROPERTY_CONTENT))
                            self.__classifier.add_potential_match(final_path)
                            return
        return

    def __check_as_cross_table(self, sheet: Worksheet, value_name_pairs: Iterator[Tuple[str, str]], path: str) -> None:
        """
        Iterates through the sheet and tries to determine if the sheet could contain a cross table where names and
        values are distributed along a column and a row and their matching is indicated by a 'X' in the field below

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        """
        # TODO: doc me
        # in a cross table everything should just stand in the table: so check if the values or the pairs can be found
        # in a column and go from there
        def check_column_for_list_entry(column: List[Cell], to_find: Iterator[str]) -> (bool, int):
            """
            Checks if the given column does contain at least one entry from the "list" given. If so true is returned
            as the index of the first match

            :param column: the column to search for a value of to_find
            :param to_find: a list of values which to be hoped to be found in the column
            :return: if a match could be found and at which index
            """
            # create a working copy
            work = list(to_find)
            start_len = float(len(work))
            earliest_hit = 1000
            for cell in column:
                current_val = str(cell)
                if current_val in work:
                    work.remove(current_val)
                    if cell.row < earliest_hit:
                        earliest_hit = cell.row
            if float(len(work)) / start_len <= 0.5:
                # if more about 50% could be found it can be assumed that it makes sense to continue
                return True, earliest_hit
            return False, -1

        def find_name_in_row(row: Iterator[Cell], to_find: str) -> str:
            """
            Goes through the given row and returns the column letter of the cell the value was found in else an
            empty string

            :param row: the row to search for the keyword
            :param to_find: the keyword to find as value in one of the cells
            :return: the column letter the value was found in
            """
            for cell in row:
                if cell.value == to_find:
                    return cell.column_letter
            return ""

        def find_x(start_index: int) -> Cell:
            """
            Returns the first cell in which a "x" could be found starting from the row index given
            :param start_index: from which index to start looking for a x
            :return: the first cell that contains a x
            """
            for row in sheet.iter_rows(min_row=start_index):
                for row_cell in row:
                    if str(row_cell) == "x" or str(row_cell) == "X":
                        return row_cell
            return None

        def find_data_field_start(to_scan: Iterator[Cell], is_row: bool) -> int:
            """
            Triggers on the first color change of cell background color that does come after a cell that was not white

            :param to_scan: the row / column to check for the color change
            :param is_row: if the iterator at hand represents a row or a column
            :return: the index of column or row the new color appeared
            """
            header_color = ""
            for cell in to_scan:
                if cell.style.bg_color == "ffffffff" and header_color == "":
                    continue    # ignore white cells which might be around the table itself
                if header_color == "":
                    header_color = cell.style.bg_color
                elif cell.style.bg_color != header_color:
                    if is_row:
                        return cell.column
                    return cell.row

        intermediate_value_name_separation = zip(*value_name_pairs)
        values = next(intermediate_value_name_separation)
        names = next(intermediate_value_name_separation)
        for col_iter in sheet.iter_cols():
            col = list(col_iter)    # transform it so we can search through it multiple times
            success_names: Tuple[bool, int] = check_column_for_list_entry(col, names)
            success_values: Tuple[bool, int] = check_column_for_list_entry(col, values)
            if not (success_names[0] and success_values[0]):
                continue
            # find the first row which contains a 'x' and try to find it's matching value in that column
            if success_names[0]:
                x_cell = find_x(success_names[1])
            else:
                x_cell = find_x(success_values[1])
            if x_cell is None:
                # does not seem to be a cross table: just abort
                return
            # find the corresponding value or name
            found_one = sheet["{}{}".format(col[0].column_letter, x_cell.row)]
            counterpart = success_values[success_names.index(found_one.value)] if success_names[0]\
                else success_names[success_values.index(found_one.value)]
            name_column = find_name_in_row(sheet[x_cell.row], counterpart)
            if name_column == "":
                # seems like it isn't a cross table after all
                return
            # derive the start of the center field from the color changes in the detected row and column
            field_start_col = find_data_field_start(sheet[x_cell.row], True)
            field_start_row = find_data_field_start(sheet[col[0].column_letter], False)
            # check which belongs where: are names in the column or values
            top_bar_pos = "{}${}".format(field_start_col, success_values[1] if success_values[0] else success_names[1])
            side_bar_pos = "${}{}".format(name_column, field_start_row)
            values_start = side_bar_pos if success_values[0] else top_bar_pos
            names_start = side_bar_pos if success_values[0] else top_bar_pos
            current_path = "{}/{}/@{}{};{};{}".format(path, sheet.title, field_start_col, field_start_row, values_start,
                                                      names_start)
            self.__classifier.add_potential_match(current_path)

    @staticmethod
    def __extract_cell_properties(to_extract_from: Cell) -> Dict[str, str]:
        """
        Takes the cell an creates a list of properties from it

        :param to_extract_from: the cell the properties are wanted from
        :return: a list of all properties supported
        """
        return {str(to_extract_from): XmlXlsxMatcher.CELL_PROPERTY_CONTENT}

    @staticmethod
    def __match_cell_properties_to(to_read_from: Cell, value_name_pairs: Iterator[Tuple[str, str]]) -> CellMatchStruct:
        """
        Extracts all supported properties of a cell (including its content) and checks if one property can be matched
        to the list of value name pairs given. This function ignores the property of the cell matched on because it
        is assumed that the first match will be on the identifier which should be the content of the cell anyway

        :param to_read_from: the cell to check the properties of
        :param value_name_pairs: a list of values expected
        :return: a struct containing data for further processing
        """
        for to_find in value_name_pairs:
            for prop in XmlXlsxMatcher.__extract_cell_properties(to_read_from).keys():
                if prop == to_find[0]:
                    # return that the value has been found
                    return XmlXlsxMatcher.CellMatchStruct(True, to_find[1], False)
                elif prop == to_find[1]:
                    # return that the name has been found
                    return XmlXlsxMatcher.CellMatchStruct(True, to_find[0], True)
        # means nothing has been found: return an invalid struct
        return XmlXlsxMatcher.CellMatchStruct(False)

    @staticmethod
    def __to_cell_address(template: str, col: str, row: int, property_identifier: str) -> str:
        """
        Inserts the given arguments into the template given. This function is merely a reminder to use a template

        :param template: the template which holds the placeholders to replace
        :param col: the column to insert
        :param row: the row to insert
        :param property_identifier: the identifier of the cell property to use as source
        :return: a properly formatted string to be used as cell address
        """
        return template.format(col, row, property_identifier)
