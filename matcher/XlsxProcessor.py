from __future__ import annotations
from typing import List, Tuple, Iterator, Dict, Any, Union
import re
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string

from matcher.internal.enum.CellPropertyType import CellPropertyType
from matcher.internal.data_struct.CellMatching import CellMatchingStruct, CellMatchResult
from matcher.internal.data_struct.CellPositioning import CellPosition, CellPositionStruct, CellPositionStructType
from classifier.BinClassifier import BinClassifier
from classifier.error.MatchExceptions import ForwardFileNotFound
from matcher.internal.data_struct.ValueNamePair import ValueNamePair
from matcher.internal.data_struct.CrossTableStruct import CrossTableStruct


class XlsxProcessor:

    FORWARDING_KEY = "forwarding_on"
    FORWARDING_PATH_KEY = "path_forward_symbol"
    TEMPLATE_CELL_ADDRESS_ROW_WISE = "${}{}:{}"
    TEMPLATE_CELL_ADDRESS_COL_WISE = "{}${}:{}"

    __classifier: BinClassifier
    __config = {}
    __root_xlsx: str
    __nested_xlsx_dir: str

    def __init__(self,
                 sink: BinClassifier,
                 config: Dict[str, str],
                 path_root_xlsx: str,
                 nested_xlsx_dir: str = "nested/"):
        """
        The constructor

        :param sink: the classifier to push the matches into
        :param config: a dictionary constructed from the config file
        :param path_root_xlsx: the path to the main Excel file
        :param nested_xlsx_dir: the path to the other Excel files which the root file might reference to
        """
        self.__classifier = sink
        self.__config = config
        self.__root_xlsx = path_root_xlsx
        root_file_name = re.search(r"\b\w*\.xlsx$", path_root_xlsx).group(0)
        root_path = path_root_xlsx[:-len(root_file_name)]
        if not nested_xlsx_dir.endswith("/"):
            nested_xlsx_dir += "/"
        self.__nested_xlsx_dir = root_path + nested_xlsx_dir

    def match_given_values_in(self, value_name_pairs: Iterator[ValueNamePair]) -> None:
        """
        Manages itself through the given xlsx-file and and tries to match the given pairs in the files (somewhere)

        :param value_name_pairs: a list of tuples with values and their corresponding URI
        """
        wb = load_workbook(self.__root_xlsx)
        for sheet in wb.sheetnames:
            self._check_row_wise(wb[sheet], value_name_pairs, self.__root_xlsx)
            self._check_column_wise(wb[sheet], value_name_pairs, self.__root_xlsx)
            self._check_as_cross_table(wb[sheet], value_name_pairs, self.__root_xlsx)

    def _check_row_wise(self, sheet: Worksheet, value_name_pairs: Iterator[ValueNamePair],
                        path: str, check_for_value_only: bool = False) -> CellPositionStruct:
        """
        Iterates through the rows of the sheet given and tries to match the given list of value-URI-pairs in a row-wise
        fashion. If a match could be made the resulting path is pushed into the classifier

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        :param check_for_value_only: set this flag to return immediately after a value from the list is found
        :return a LineResultStruct if only a value has been found: this is important for forwarded searches else an
                invalid / empty struct
        """
        current_sheet_path = "{}/{}".format(path, sheet.title)
        lowest_header_row: int = 0
        forward_index = -1
        handle_forwarding, forwarding_column_name = self.__includes_forwarding(sheet.title)
        color_key = "header_{}".format(sheet.title)
        if color_key not in self.__config:
            raise AssertionError("Missing key '{}' for retrieving the header color".format(color_key))
        header_color = self.__config[color_key]
        row_index = 0   # start with zero to be conform with xlsx
        for row in sheet.iter_rows():
            row_index += 1
            result = self.__scan_cell_line_for(row, current_sheet_path, value_name_pairs, forward_index, header_color,
                                               "" if not handle_forwarding else forwarding_column_name)
            if result.read_result == CellPositionStructType.NO_FINDING:
                continue
            elif result.read_result == CellPositionStructType.HEADER_FOUND:
                lowest_header_row = row_index
                if result.contains_header_forwarding_position():
                    forward_index = column_index_from_string(result.name_or_forward_position.column)
                continue
            # else data has been found
            if result.match_struct.success_type == CellMatchResult.VALUE_FOUND and check_for_value_only:
                # only a value has been found -> forward the information to the caller -> but if no header has been
                # found the orientation is probably bogus
                if lowest_header_row > 0:
                    return result
            elif result.match_struct.success_type != CellMatchResult.ALL_FOUND:
                # this case shouldn't exist in reality yet cover it to be on the safe side
                continue
            # else a pair has been detected -> collect all the required data and push the result into the classifier
            row_data_start = lowest_header_row + 1
            value_cell = self.__to_linear_cell_address(False, result.value_position.column, row_data_start,
                                                       result.value_position.read_type)
            value_path = current_sheet_path if not result.contains_value_forwarding_path() else result.value_path
            final_path = "{}/@{};".format(value_path, value_cell)
            name_cell = self.__to_linear_cell_address(False, result.name_or_forward_position.column, row_data_start,
                                                      result.name_or_forward_position.read_type)
            if result.contains_value_forwarding_path():
                # prepend the current path to the name path to indicate that both differ
                name_cell = "{}/@{}".format(current_sheet_path, name_cell)
            else:
                name_cell = "@{}".format(name_cell)
            final_path += name_cell
            self.__classifier.add_potential_match(final_path)
        # if something was found it has been pushed to the classifier already -> so no need to return anything
        return CellPositionStruct.create_no_find()

    def _check_column_wise(self, sheet: Worksheet, value_name_pairs: Iterator[ValueNamePair],
                           path: str, check_for_value_only: bool = False) -> CellPositionStruct:
        """
        Iterates through the columns of the sheet given and tries to match the given list of value-URI-pairs in a
        column-wise fashion. If a match could be made the resulting path is pushed into the classifier

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        :param check_for_value_only: set this flag to return immediately after a value from the list is found
        :return a LineResultStruct if only a value has been found: this is important for forwarded searches else an
                invalid / empty struct
        """
        current_sheet_path = "{}/{}".format(path, sheet.title)
        lowest_header_col: int = 0
        forward_index = -1
        handle_forwarding, forwarding_row_name = self.__includes_forwarding(sheet.title)
        header_color = self.__config["header_{}".format(sheet.title)]
        col_index = -1
        for column in sheet.iter_cols():
            col_index += 1
            result = self.__scan_cell_line_for(column, current_sheet_path, value_name_pairs, forward_index,
                                               header_color, "" if not handle_forwarding else forwarding_row_name)
            if result.read_result == CellPositionStructType.NO_FINDING:
                continue
            elif result.read_result == CellPositionStructType.HEADER_FOUND:
                lowest_header_col = col_index
                if result.contains_header_forwarding_position():
                    forward_index = result.name_or_forward_position.row
                continue
            # else data has been found
            if result.match_struct.success_type == CellMatchResult.VALUE_FOUND and check_for_value_only:
                # only a value has been found -> forward the information to the caller -> but if no header has been
                # found the orientation is probably bogus
                if lowest_header_col > 0:
                    return result
            elif result.match_struct.success_type != CellMatchResult.ALL_FOUND:
                # this case shouldn't exist in reality yet cover it to be on the safe side
                continue
            # else a pair has been detected -> collect all the required data and push the result into the classifier
            col_data_start = get_column_letter(lowest_header_col + 1)
            value_cell = self.__to_linear_cell_address(True, col_data_start, result.value_position.row,
                                                       result.value_position.read_type)
            value_path = current_sheet_path if not result.contains_value_forwarding_path() else result.value_path
            final_path = "{}/@{};".format(value_path, value_cell)
            name_cell = self.__to_linear_cell_address(True, col_data_start, result.name_or_forward_position.row,
                                                      result.name_or_forward_position.read_type)
            if result.contains_value_forwarding_path():
                # prepend the current path to the name path to indicate that both differ
                name_cell = "{}/@{}".format(current_sheet_path, name_cell)
            else:
                name_cell = "@{}".format(name_cell)
            final_path += name_cell
            self.__classifier.add_potential_match(final_path)
        # if something was found it has been pushed to the classifier already -> so no need to return anything
        return CellPositionStruct.create_no_find()

    def _check_as_cross_table(self, sheet: Worksheet, value_name_pairs: Iterator[ValueNamePair], path: str) -> None:
        """
        Iterates through the sheet and tries to determine if the sheet could contain a cross table where names and
        values are distributed along a column and a row and their matching is indicated by a 'X' in the field below

        :param sheet: the sheet to search for matches
        :param value_name_pairs: the stuff (hopefully) to find in the sheet given
        :param path: the path to the current sheet (excluding the sheet itself)
        """
        # in a cross table everything should just stand in the table: so check if the values or the pairs can be found
        # in a column and go from there
        def check_line_for_list_entries(line: List[Cell], to_find: Iterator[str]) -> (bool, int):
            """
            Checks if the given column or row does contain at least 50% of the entries from the "list" given. If so true
            is returned as the index of the first match

            :param line: the column or row to search for a value of to_find
            :param to_find: a list of values which to be hoped to be found in the column
            :return: if a match could be found and at which index
            """
            # create a working copy
            work = list(to_find)
            start_len = float(len(work))
            earliest_hit = 1000
            for cell in line:
                if cell.value is None:
                    continue
                current_val = cell.value
                if current_val in work:
                    work.remove(current_val)
                    if cell.row < earliest_hit:
                        earliest_hit = cell.row
            if float(len(work)) / start_len <= 0.5:
                # if more about 50% could be found it can be assumed that it makes sense to continue
                return True, earliest_hit
            return False, -1

        def find_name_in_col(column: Iterator[Cell], to_find: str) -> CellPosition:
            """
            Goes through the given column and returns the cell the value was found in else None

            :param column: the column to search for the keyword
            :param to_find: the keyword to find as value in one of the cells
            :return: the row index the value was found in
            """
            for cell in column:
                if cell.value == to_find:
                    return CellPosition.create_from(cell)
            return CellPosition.create_invalid()

        def find_x(start_index: int) -> CellPosition:
            """
            Returns the first cell in which a "x" could be found starting from the row index given
            :param start_index: from which index to start looking for a x
            :return: the first cell that contains a x
            """
            for row in sheet.iter_rows(min_row=start_index):
                row_cell: Cell
                for row_cell in row:
                    if row_cell.value == "x" or row_cell.value == "X":
                        return CellPosition.create_from(row_cell)
            return CellPosition.create_invalid()

        def find_xs(start_row: int, how_many: int) -> bool:
            """
            Goes through the rows below the given one and counts every single X it encounters: if the count reaches the
            how_many-argument true is returned

            :param start_row: the row to start searching at
            :param how_many: the many x's should be found
            :return: true if the expected count was reached else false
            """
            x_count = 0
            for row in sheet.iter_rows(min_row=start_row):
                for cell in row:
                    if cell.value == "x" or cell.value == "X":
                        x_count += 1
                    if x_count >= how_many:
                        return True
            return False

        def scan_rows_for_list_until(data: CrossTableStruct) -> Tuple[bool, CrossTableStruct]:
            """
            Goes through the rows of the sheet in search for the list of the other part of the value-name pairs until
            the given row. Returns the given struct again to indicate that the struct might be altered

            :param data: the struct that holds all the required data
            :return: a tuple if enough values could be found and the updated struct in this case
            """
            for row in sheet.iter_rows(max_row=data.first_find.row):
                found_list = data.find_other_pair_values_in(row)
                if found_list:
                    # the second position had been stored with the check already
                    return True, data
            return False, data


        def find_data_field_start(to_scan: Iterator[Cell]) -> CellPosition:
            """
            Triggers on the first color change of cell background color that does come after a cell that was not white

            :param to_scan: the row / column to check for the color change
            :return: the position at which the transition from non-white to another color appeared
            """
            header_color = ""
            for cell in to_scan:
                if header_color == "" and cell.fill.bgColor.rgb == "00000000":
                    continue    # ignore white cells which might be around the table itself
                elif cell.fill.bgColor.rgb != header_color:
                    if header_color == "":
                        header_color = cell.fill.bgColor.rgb
                    else:
                        # color has already been set so there's a new color at hand -> table header is reached
                        return CellPosition.create_from(cell)
            return CellPosition.create_invalid()

        def value_of(position: CellPosition) -> str:
            """
            Returns the cell content at the given position

            :param position: where to extract the value from
            :return: the value of the cell
            """
            return sheet["{}{}".format(position.column, position.row)].value

        values, names = ValueNamePair.unzip(value_name_pairs)
        for col in sheet.iter_cols():
            success, progress_struct = CrossTableStruct.values_exist_in(col, value_name_pairs)
            if not success:
                continue
            # a cross matrix should have a side and a top -> now check if the top exists: it should be above the first
            # value
            success, progress_struct = scan_rows_for_list_until(progress_struct)
            if not success:
                # probably not a cross table -> stop here
                return
            # now check if we find some "x" below which is a strong indicator in combination with the results above
            success = find_xs(progress_struct.opposite_find.row,
                              CrossTableStruct.REQUIRED_SUCCESS_RATE * progress_struct.get_sample_size())
            if not success:
                # no cross table after all
                return
            # if this line has been reached it is safe to assume to have a cross table at hand
            # TODO: check for the colors now

            # # find the first row which contains a 'x' and try to find it's matching value in that column
            # if success_names:
            #     x_cell = find_x(name_index)
            # else:
            #     x_cell = find_x(value_index)
            # if not x_cell.is_valid():
            #     # does not seem to be a cross table: just abort
            #     return
            # # TODO: a proper implementation might better check for multiple 'X'
            # # find the corresponding value or name
            # found_one = sheet["{}{}".format(, x_cell.row)]
            # counterpart_content = values[names.index(found_one.value)] if success_names else names[values.index(
            #     found_one.value)]
            # counterpart_pos = find_name_in_col(sheet[x_cell.column], counterpart_content)
            # if counterpart_pos is None:
            #     # seems like it isn't a cross table after all
            #     return
            # # perform a final check if the column contains eg. most of the names the row should contain most of the
            # # values
            # confirmed, _ = check_line_for_list_entries(sheet[counterpart_pos.row], names if success_values else values)
            # if not confirmed:
            #     # most of the expected counter values could not be found in a row -> doesn't seem to be cross table
            #     return
            # # derive the start of the center field from the color changes in the detected row and column -> this is
            # # should be more robust then working with the names itself which might be incomplete
            # field_start_col = get_column_letter(find_data_field_start(sheet[x_cell.row], True))
            # field_start_row = find_data_field_start(sheet[counterpart_pos.column], False)
            # # check which belongs where: are names in the column or values
            # top_bar_pos = "{}${}".format(field_start_col, counterpart_pos.row)
            # side_bar_pos = "${}{}".format(col[0].column_letter, field_start_row)
            # values_start = side_bar_pos if success_values else top_bar_pos
            # names_start = top_bar_pos if success_values else side_bar_pos
            # current_path = "{}/{}/@{}{};{};{}".format(path, sheet.title, field_start_col, field_start_row, values_start,
            #                                           names_start)
            # self.__classifier.add_potential_match(current_path)
            # # the cross table has been identified, next attempts should fail anyway so abort search
            return

    def _follow_forward_to(self, file_name: str, work_path: str,
                           testing_struct: CellMatchingStruct) -> CellPositionStruct:
        """
        Performs a row- and a column-wise search for the value in the file specified

        :param file_name: the file name to check for the value
        :param work_path: the internal path which describes the "address" from the root-file
        :param testing_struct: the struct which can be used to find the value
        :return: A CellPositionStruct of "type" value_found or no_find
        """
        if testing_struct.success_type == CellMatchResult.NO_FINDING:
            # it makes no sense to look in another file without a previously found name
            return CellPositionStruct.create_no_find()
        file_path = self.__nested_xlsx_dir + file_name
        if not os.path.exists(file_path):
            raise ForwardFileNotFound("Could not find {} file under: {}".format(file_name, file_path))
        path = work_path + "/{}".format(self.__config[self.FORWARDING_PATH_KEY])
        # create a dummy list which only contains the missing entry -> which has to be value else the forwarding would
        # be stupid
        value_pair: Iterator[ValueNamePair] = [ValueNamePair.create_with_value(testing_struct.get_missing_entry())]
        wb = load_workbook(file_path)
        for sheet in wb.sheetnames:
            # return the first value found
            result_row = self._check_row_wise(wb[sheet], value_pair, path, True)
            if result_row.contains_value_forwarding_path():
                return result_row
            result_col = self._check_column_wise(wb[sheet], value_pair, path, True)
            if result_col.contains_value_forwarding_path():
                return result_col
        return CellPositionStruct.create_no_find()

    def __includes_forwarding(self, sheet_name: str) -> Tuple[bool, str]:
        """
        Checks if the given sheet name is registered with a column which forwards to another file

        :param sheet_name: the name to check for
        :return: a tuple which contains if forwarding is to be expected and the column which is to be used for that
        """
        # as this project has only one root file the separation is simple
        names = self.__config[self.FORWARDING_KEY].split('/')
        if len(names) != 2:
            raise ValueError("The config file is wrong: expecting a path made of sheet/column. Received: " +
                             self.__config[self.FORWARDING_KEY])
        if names[0] != sheet_name:
            return False, ""
        return True, names[1]

    def __scan_cell_line_for(self, to_scan: Iterator[Cell], current_path: str,
                             value_name_pairs: Iterator[ValueNamePair] = None, forward_index: int = -1,
                             header_color: str = "", forward_name: str = "") -> CellPositionStruct:
        """
        Goes through the iteration of cells and checks if it can find useful information in it. The result is returned
        in form of struct which holds data depending on the found data

        :param to_scan: the iterator for a collection of cells (which is usually either a column or a row)
        :param current_path: the current path within the file / sheet structure
        :param value_name_pairs: the values and names to find in the sheet. Set if you want to find these data
        :param forward_index: set this parameter if at this given index the content has to be interpreted as file name
        :param header_color: set this value if a header has to be detected
        :param forward_name: set this value if a header field containing this value indicates forwarding
        :return: a situation dependent initialized instance of LineResultStruct
        """
        def has_header_color(to_check: Cell) -> bool:
            """
            Checks if the given cell has a background color equal to the headers background color

            :param to_check: the cell to check for coloring
            :return: true if the colors match else false
            """
            return to_check.fill.bgColor.rgb == header_color or to_check.fill.fgColor.rgb == header_color

        if value_name_pairs is None:
            value_name_pairs = []
        current_idx = 0     # which results in starting the iteration with 1 which is the start for excel
        value_path = ""
        result_struct = CellMatchingStruct(value_name_pairs)
        value_position = CellPosition.create_invalid()
        name_position = CellPosition.create_invalid()
        is_header = False
        for cell in to_scan:
            current_idx += 1
            if cell.value is None:
                # skip the empty cell
                continue
            if has_header_color(cell):
                if cell.value == forward_name:
                    # return immediately as only one forwarding index is expected
                    # -> if multiple are required use a state machine
                    return CellPositionStruct.create_header_found(CellPosition.create_from(cell,
                                                                                           CellPropertyType.CONTENT))
                is_header = True
                continue
            # now the cell should contain some data -> find out if it is data of interest
            if current_idx == forward_index:
                result = self._follow_forward_to(cell.value, current_path, result_struct)
                # write directly to value_position as the name shouldn't be found in the forwarding file
                # -> this would beat the whole purpose of the forwarding
                result_struct = result.match_struct
                value_position = result.value_position
                value_path = result.value_path
            else:
                cell_data = XlsxProcessor.__extract_cell_properties(cell)
                for data in cell_data.keys():
                    result = result_struct.test_value(data)
                    if result == CellMatchResult.NAME_FOUND:
                        name_position = CellPosition.create_from(cell, cell_data[data])
                    elif result == CellMatchResult.VALUE_FOUND:
                        value_position = CellPosition.create_from(cell, cell_data[data])
            if name_position.is_valid() and value_position.is_valid():
                # no reason to continue -> everything has been found
                return CellPositionStruct.create_data_pair_found(result_struct, value_position, name_position)
        # keep this separated as their separation ensures that header and data are detected in different lines
        # -> prefer the value over the header as no detected header means no correctly detected table
        if value_position.is_valid():
            return CellPositionStruct.create_value_found(result_struct, value_position, value_path)
        if is_header:
            return CellPositionStruct.create_header_found(CellPosition.create_invalid())
        return CellPositionStruct.create_no_find()

    @staticmethod
    def __extract_cell_properties(to_extract_from: Cell) -> Dict[str, CellPropertyType]:
        """
        Takes the cell an creates a list of properties from it

        :param to_extract_from: the cell the properties are wanted from
        :return: a list of all properties supported
        """
        return {str(to_extract_from.value): CellPropertyType.CONTENT}

    @staticmethod
    def __to_linear_cell_address(is_fixed_row: bool, col: str, row: int, property_identifier: CellPropertyType) -> str:
        """
        Inserts the given arguments into the template given. This function is merely a reminder to use a template

        :param is_fixed_row: if the template should assume a row-wise reading scheme or a column-wise scheme
        :param col: the column to insert
        :param row: the row to insert
        :param property_identifier: the identifier of the cell property to use as source
        :return: a properly formatted string to be used as cell address
        """
        if is_fixed_row:
            template = XlsxProcessor.TEMPLATE_CELL_ADDRESS_COL_WISE
        else:
            template = XlsxProcessor.TEMPLATE_CELL_ADDRESS_ROW_WISE
        return template.format(col, row, str(property_identifier))
