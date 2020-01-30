import random
from openpyxl import load_workbook

from creation.WorkerCreator import WorkerCreator
from creation.NameProvider import NameProvider


class XlsxModifier:

    first_row_workers = 7
    first_row_affiliations = 5

    @staticmethod
    def update_worker_xlsx(worker_cnt: int, source_file: str) -> None:
        """
        Changes the age of a randomly picked worker and the size of another randomly picked worker

        :param worker_cnt: the number of workers to pick from
        :param source_file: the file to modify
        """
        print("Assuming C{} in Workers as first entry in name column".format(XlsxModifier.first_row_workers))
        print("Assuming columns E, F & G represent the Size, the Age & the Birthday in Workers")
        wb = load_workbook(source_file)
        worker_sheet = wb["Workers"]
        first_victim = random.randint(0, worker_cnt - 1)
        new_age = 67
        print("Changing the age of {} to {}".format(worker_sheet["C{}".format(
            XlsxModifier.first_row_workers + first_victim)].value, new_age))
        worker_sheet["F{}".format(XlsxModifier.first_row_workers + first_victim)].value = str(new_age)
        new_birthday = WorkerCreator.draw_birthday(new_age)
        worker_sheet["G{}".format(XlsxModifier.first_row_workers + first_victim)].value = WorkerCreator.format_date(new_birthday)
        second_victim = random.randint(0, worker_cnt - 1)
        new_size = "XS"
        print("Changing size of {} to {}. This should have no impact on the final XML-file".format(
            worker_sheet["C{}".format(XlsxModifier.first_row_workers + second_victim)].value, new_size))
        worker_sheet["E{}".format(XlsxModifier.first_row_workers + second_victim)].value = new_size
        wb.save(source_file)

    @staticmethod
    def add_worker_to_xlsx(worker_cnt: int, source_file: str) -> None:
        """
        Adds a new worker to the given xlsx-file

        :param worker_cnt: the count of existing workers
        :param source_file: the file to modify
        """
        # assume the same cells like in the update function
        new_name = NameProvider.get_name()
        new_id = str(100000)
        new_job = WorkerCreator.job_controller
        new_size = "S"
        new_age = random.randint(48, 59)
        new_birthday = WorkerCreator.format_date(WorkerCreator.draw_birthday(new_age))
        print("Adding {} as {} to the list of workers".format(new_name, new_job))
        wb = load_workbook(source_file)
        worker_sheet = wb["Workers"]
        line = str(XlsxModifier.first_row_workers + worker_cnt)
        worker_sheet["B" + line].value = new_id
        worker_sheet["C" + line].value = new_name
        worker_sheet["D" + line].value = new_job
        worker_sheet["E" + line].value = new_size
        worker_sheet["F" + line].value = str(new_age)
        worker_sheet["G" + line].value = new_birthday

        affiliations_sheet = wb["Affiliations"]
        affiliations_sheet["B{}".format(XlsxModifier.first_row_affiliations + worker_cnt)].value = new_name
        # make it management
        affiliations_sheet["G{}".format(XlsxModifier.first_row_affiliations + worker_cnt)].value = "X"
        wb.save(source_file)
