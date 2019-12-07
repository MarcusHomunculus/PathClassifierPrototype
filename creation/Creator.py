from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from typing import Dict, List, Tuple
from datetime import date
import json
import re
import random
import math
import os
import xml.etree.ElementTree as ElemTree
from xml.dom import minidom
from creation.SectionCreator import SectionCreator


class InitializationException(Exception):
    """
    Exception is intended for situations when a class should have been provided
    with data before but wasn't
    """


class Creator:
    """
    The class which creates the files which the learning algorithm can use for
    training
    """

    # constants
    WORKER_SHEET = "Workers"
    SECTION_SHEET = "Sections"

    class DataStruct:
        """
        Works as a container / struct to organize the data for a worker
        """
        attributes: Dict[str, str]
        skills: List[str]

        def __init__(self):
            self.attributes = {}
            self.skills = []

        def __str__(self):
            s = "Content:\n"
            for attr in self.attributes:
                s += "{} : {}\n".format(attr, self.attributes[attr])
            return s + "List content: {}".format(str(self.skills))

    __sectionList: List[DataStruct] = []
    __workerList: List[DataStruct] = []
    __assignments: List[Tuple[DataStruct, DataStruct]] = []

    def __init__(self, path_to_workers: str, path_to_sections: str):
        """
        The constructor which is a wrapper around read_from_json()

        :param path_to_workers: the path to the file containing the worker
                                definitions in JSON formatting
        :param path_to_sections: the path to the file containing the section
                                 definitions in JSON formatting
        """
        self.read_from_json(path_to_workers, path_to_sections)

    def read_from_json(self, path_to_workers: str, path_to_sections: str) -> None:
        """
        Reads in the data from the 2 JSON-files in order to create data from
        them
        :param path_to_workers: the path to the file containing the worker
                                definitions in JSON formatting
        :param path_to_sections: the path to the file containing the section
                                 definitions in JSON formatting
        """
        def read_in(path: str, target: List[Creator.DataStruct]) -> None:
            """
            Writes the data from the specified JSON-file into the list specified, too
            :param path: the path to the JSON file
            :param target: the list in which to dump the converted data
            """
            with open(path) as json_file:
                entities = json.load(json_file)
                for e in entities:
                    current = Creator.DataStruct()
                    for attr in e:
                        if attr == "Skills":
                            current.skills = e[attr]
                            continue
                        current.attributes[attr] = e[attr]
                    target.append(current)

        for t in ["workers", "sections"]:
            if t == "workers":
                read_in(path_to_workers, self.__workerList)
            elif t == "sections":
                read_in(path_to_sections, self.__sectionList)
            else:
                raise AttributeError(
                    "No implementation to push {} to the internal members".format(t))
        # continue with matching sections with workers
        self.__assign()

    def create_xlsx(self, target_dir: str = "../", main_file_name: str = "data.xlsx",
                    section_dir: str = "sections") -> None:
        """
        Generates the files which represent the data in a xlsx structure

        :param target_dir: the path to the directory were the files should be dumped into
        :param main_file_name: the name of the main data file
        :param section_dir: an additional path to the target dir to "nest" the section files into
        """
        if not self._has_internal_data():
            raise InitializationException("Members seem not been initialized with data")
        # check if all paths end with a separator
        if not target_dir.endswith("/"):
            target_dir += "/"
        if not section_dir.endswith("/"):
            section_dir += "/"
        # check if the main file name already contains the correct file extension
        if not main_file_name.endswith(".xlsx"):
            main_file_name += "xlsx"
        # create the sheet with the worker data
        wb = Workbook()
        # a blank sheet is created by default: use this instead of creating a new one
        current = wb.active
        current.title = self.WORKER_SHEET
        # foreach property of worker
        self._create_single_table(current, self.__workerList, 4, 2, self.WORKER_SHEET)
        # create the sheet with the sections
        current = wb.create_sheet(self.SECTION_SHEET)
        self._create_single_table(current, self.__sectionList, 4, 2, self.SECTION_SHEET)
        # create the sheet where to every skill of a section is corresponded with a workers skill
        current = wb.create_sheet("Affiliations")
        self._create_cross_table_simple(current, 2, 2, "Affiliations")
        wb.save(target_dir + main_file_name)
        wb.close()
        # continue with the referenced files
        if not os.path.exists(target_dir + section_dir):
            os.mkdir(target_dir + section_dir)
        self._create_team_xlsx_file(target_dir + section_dir + "{name}.xlsx", 2, 2)

    def create_xml(self, file_name: str) -> None:
        def prettify(elem: ElemTree.ElementTree) -> str:
            """
            Corrects the indentation for the generated XML
            :param elem: the final tree to prettify
            :return: a string which has the well known XML-structure
            """
            # courtesy goes to https://pymotw.com/2/xml/etree/ElementTree/create.html#pretty-printing-xml
            tree_str = ElemTree.tostring(elem.getroot(), encoding='utf-8', method='xml')
            restructured = minidom.parseString(tree_str)
            return restructured.toprettyxml(indent="  ")

        if not self._has_internal_data():
            raise InitializationException("Members seem not been initialized with data")
        root = ElemTree.Element("company")
        meta = ElemTree.SubElement(root, "general")
        name = ElemTree.SubElement(meta, "name")
        name.text = "The Product Company"
        founded = ElemTree.SubElement(meta, "founded")
        founded.text = str(date.today())
        trade = ElemTree.SubElement(meta, "trade")
        trade.text = "Product production"
        struct = ElemTree.SubElement(root, "company_structure")
        departments = ElemTree.SubElement(struct, "sections")
        for depart in self.__sectionList:
            current_section = ElemTree.SubElement(departments, "section")
            section_name = ElemTree.SubElement(current_section, "name")
            section_name.text = depart.attributes["Name"]
            worker_factor = ElemTree.SubElement(current_section, "normalized_worker_count",
                                                {"PaymentStage": str(depart.attributes["PaymentStage"])})
            worker_factor.text = str(depart.attributes["NormalizedWorkerCount"])
            section_teams = ElemTree.SubElement(current_section, "teams")
            for team in depart.attributes["Teams"]:
                current_team = ElemTree.SubElement(section_teams, "team", {"name": team})
                current_team.text = str(self.__team_size_from_workers_and_team_fraction(
                    float(depart.attributes["NormalizedWorkerCount"]), float(depart.attributes["Teams"][team])))
            assigned_workers = ElemTree.SubElement(current_section, "workers")
            for assignment in self.__assignments:
                if assignment[0] == depart:
                    worker = assignment[1]
                    current_worker = ElemTree.SubElement(assigned_workers, "worker")
                    current_worker.text = worker.attributes["Name"]
        tree = ElemTree.ElementTree(root)
        with open(file_name, "w") as file:
            print(prettify(tree), file=file)

    def _has_internal_data(self) -> bool:
        """
        Checks if the members are filled with data else false is returned
        :return: true if all members hold data and are not empty lists
        """
        return len(self.__sectionList) != 0 and len(self.__workerList) != 0

    def _create_single_table(self, workbook, data: List[DataStruct], start_row: int, start_column: int, name: str,
                             offset_row: int = 2, offset_col: int = 0) -> None:
        """
        Creates a row table dedicated to the data given
        :param workbook: the to write the data in
        :param data: the actual data to write (should either be worker or section descriptions)
        :param start_row: the row the table heading should be placed in
        :param start_column: the column the table heading should start
        :param name: the name of the table to use as heading
        :param offset_row: set this value for vertical displacement between the heading and the table header
        :param offset_col: set this value for horizontal displacement between the heading the the table header
        """
        workbook.cell(row=start_row, column=start_column).value = name
        # use the first data point as template
        header_keys = data[0].attributes.keys()
        # create the table header
        header_fill = PatternFill(start_color='FFFFFF00',
                                  end_color='FFFFFF00', fill_type='solid')
        header_alignment = Alignment(horizontal='center')
        table_row = start_row + offset_row
        current_col = start_column + offset_col
        for key in header_keys:
            current_cell = workbook.cell(row=table_row, column=current_col)
            key_content = key if key != SectionCreator.TEAM_KEY else "section file"
            current_cell.value = key_content
            # do some styling
            current_cell.fill = header_fill
            current_cell.alignment = header_alignment
            # adapt the width in the process
            col_letter = get_column_letter(current_col)
            # use the first worker as a template again
            # take whatever is longer: key or content
            content_width = len(str(data[0].attributes[key]))
            key_width = len(key)
            width = key_width if key_width > content_width else content_width
            workbook.column_dimensions[col_letter].width = 1.8 * width
            current_col += 1
        # update the row "pointer" below the header
        table_row += 1
        for y in range(len(data)):
            current_col = start_column + offset_col
            for key in header_keys:
                cell = workbook.cell(row=table_row + y, column=current_col)
                val = data[y].attributes[key] if key != SectionCreator.TEAM_KEY else Creator._replace_spaces(
                    data[y].attributes["Name"]) + ".xlsx"
                cell.value = self.__val_to_string(val)
                current_col += 1

    def _create_cross_table(self, workbook, start_row: int, start_column: int, table_tile: str, offset_row: int = 2,
                            offset_col: int = 0):
        """
        Creates a matrix which shows which worker fits to which section based on the skills

        :param workbook: the sheet to write the table in
        :param start_row: the row the heading should be placed in
        :param start_column: the column the heading should be placed in
        :param table_tile: the title to use for the heading
        :param offset_row: the displacement between the heading and the table itself in row count
        :param offset_col: the displacement between the heading the the table itself in columns
        """
        # start with creating the header
        workbook.cell(row=start_row, column=start_column).value = table_tile
        current_row = start_row + offset_row + 1    # start one row below and merge all section cells afterwards
        # set row height for the section skills based on the first entry using it as blueprint for length estimation
        workbook.row_dimensions[current_row].height = len(self.__sectionList[0].skills[0]) * 5 * 1.7
        # and for the names
        workbook.row_dimensions[current_row-1].height = len(self.__sectionList[0].attributes["Name"]) * 10
        # +2 for the col as downwards the columns for the worker and the abilities are required
        current_column = start_column + offset_col + 2
        # set the styling
        header_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                               bottom=Side(style='medium'))
        section_fill = PatternFill(start_color='FFFF9933', end_color='FFFF9933', fill_type='solid')
        for section in self.__sectionList:
            col_section_start = current_column
            for skill in section.skills:
                # adapt the column width where required
                workbook.column_dimensions[get_column_letter(current_column)].width = 3
                active = workbook.cell(row=current_row, column=current_column)
                active.value = skill
                active.alignment = Alignment(text_rotation=90)
                active.fill = section_fill
                active.border = header_border
                current_column += 1
            # now merge the upper cells and insert the section name
            if len(section.skills) > 1:
                workbook.merge_cells(start_row=current_row-1, start_column=col_section_start, end_row=current_row-1,
                                     end_column=current_column-1)
            section_cell = workbook.cell(row=current_row-1, column=col_section_start)
            section_cell.value = section.attributes["Name"]
            section_cell.alignment = Alignment(text_rotation=90, horizontal='center')
            section_cell.fill = section_fill
            section_cell.border = header_border
        # continue with the workers: fill with the skills and merge later with the name one column before
        current_row += 1
        skill_col = start_column + offset_col + 1
        workbook.column_dimensions[get_column_letter(skill_col - 1)].width = 20
        workbook.column_dimensions[get_column_letter(skill_col)].width = 20
        worker_fill = PatternFill(start_color='FF66FF33', end_color='FF66FF33', fill_type='solid')
        for worker in self.__workerList:
            row_worker_start = current_row
            for skill in worker.skills:
                active = workbook.cell(row=current_row, column=skill_col)
                active.value = skill
                active.fill = worker_fill
                active.border = header_border
                current_row += 1
            workbook.merge_cells(start_row=row_worker_start, start_column=skill_col-1, end_row=current_row-1,
                                 end_column=skill_col-1)
            worker_cell = workbook.cell(row=row_worker_start, column=skill_col-1)
            worker_cell.value = worker.attributes["Name"]
            worker_cell.alignment = Alignment(vertical='center')
            worker_cell.fill = worker_fill
            worker_cell.border = header_border
        # cross_space_fill = PatternFill(start_color='FFFFFF66', end_color='FFFFFF66', fill_type='solid')
        # cross_space_border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        raise NotImplementedError("Function is not fully implemented")

    def _create_cross_table_simple(self, workbook, start_row: int, start_column: int, table_tile: str,
                                   offset_row: int = 2, offset_col: int = 0) -> None:
        """
        Creates the table which shows which employee is assigned to which department
        :param workbook: the sheet to write the table in
        :param start_row: the row the heading should be placed in
        :param start_column: the column the heading should be placed in
        :param table_tile: the title to use for the heading
        :param offset_row: the displacement between the heading and the table itself in row count
        :param offset_col: the displacement between the heading the the table itself in columns
        """
        # start with creating the header
        workbook.cell(row=start_row, column=start_column).value = table_tile
        # set the styling
        header_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                               bottom=Side(style='medium'))
        section_fill = PatternFill(start_color='FFFF9933', end_color='FFFF9933', fill_type='solid')
        current_row = start_row + offset_row
        # set row height for the section names based on the first entry using it as blueprint for length estimation
        workbook.row_dimensions[current_row].height = len(self.__sectionList[0].attributes["Name"]) * 12
        # +1 for the col as downwards the columns for the worker and the abilities are required
        current_column = start_column + offset_col + 1
        for section in self.__sectionList:
            workbook.column_dimensions[get_column_letter(current_column)].width = 3
            active = workbook.cell(row=current_row, column=current_column)
            active.value = section.attributes["Name"]
            active.alignment = Alignment(text_rotation=90)
            active.fill = section_fill
            active.border = header_border
            current_column += 1
        # continue with the workers
        current_row += 1
        current_column = start_column + offset_col
        workbook.column_dimensions[get_column_letter(current_column)].width = 20
        worker_fill = PatternFill(start_color='FF66FF33', end_color='FF66FF33', fill_type='solid')
        for worker in self.__workerList:
            active = workbook.cell(row=current_row, column=current_column)
            active.value = worker.attributes["Name"]
            active.fill = worker_fill
            active.border = header_border
            current_row += 1
        # now insert the matching data
        cross_space_fill = PatternFill(start_color='FFFFFF66', end_color='FFFFFF66', fill_type='solid')
        cross_space_border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        field_row = start_row + offset_row + 1
        field_column = start_column + offset_col + 1
        # fill the whole space with color and borders
        for x in range(len(self.__sectionList)):
            for y in range(len(self.__workerList)):
                active = workbook.cell(row=field_row + y, column=field_column + x)
                active.border = cross_space_border
                active.fill = cross_space_fill
                active.alignment = Alignment(horizontal='center')
        for pair in self.__assignments:
            current_column = self.__sectionList.index(pair[0]) + field_column
            current_row = self.__workerList.index(pair[1]) + field_row
            active = workbook.cell(row=current_row, column=current_column)
            active.value = "X"

    def _create_team_xlsx_file(self, path: str, start_row: int, start_column: int,
                               offset_row: int = 2, offset_col: int = 1) -> None:
        """
        Creates the file which represents the allowed team distribution within the given department

        :param path: the path under which the file is to be created
        :param start_row: the row of the heading
        :param start_column: the column to write the heading in
        :param offset_row: the displacement of the "header" to the heading in y direction
        :param offset_col: the displacement of the "header" to the heading in x direction
        """
        for section in self.__sectionList:
            wb = Workbook()
            current = wb.active
            team_attr = "Teams"
            section_name = section.attributes["Name"]
            current.title = section_name
            current.cell(row=start_row, column=start_column).value = "Maximal allowed team sizes in {}".format(
                section_name)
            current_row = start_row + offset_row
            current_column = start_column + offset_col
            # transform the data into a more suitable format
            team_dict = {}
            biggest_team = 0
            for team in section.attributes[team_attr]:
                head_count = self.__team_size_from_workers_and_team_fraction(
                    float(section.attributes["NormalizedWorkerCount"]), float(section.attributes[team_attr][team]))
                team_dict[team] = head_count
                if head_count > biggest_team:
                    biggest_team = head_count
            header_fill = PatternFill(start_color='FFFFFF66', end_color='FFFFFF66', fill_type='solid')
            border = Border(right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                   top=Side(style='thin'))
            for i in range(biggest_team):
                active = current.cell(row=current_row, column=current_column + i)
                active.value = i + 1
                active.fill = header_fill
                active.border = border
                active.alignment = Alignment(horizontal='center')
            current_row += 1
            team_fill = PatternFill(start_color='FF66FF99', fill_type='solid')
            for team in team_dict.keys():
                current.merge_cells("{col_1}{row}:{col_2}{row}".format(row=current_row,
                                                                       col_1=get_column_letter(current_column),
                                                                       col_2=get_column_letter(current_column +
                                                                                               team_dict[team] - 1)))
                active = current.cell(row=current_row, column=current_column)
                active.value = "{0} : {1}".format(team, team_dict[team])
                active.fill = team_fill
                active.border = border
                current_row += 1
            file_name = Creator._replace_spaces(section_name)
            wb.save(path.format(name=file_name))
            wb.close()

    @staticmethod
    def _replace_spaces(to_transform: str) -> str:
        """
        Replaces all whitespaces with underscores in the string given
        """
        return re.sub(r"\s", "_", to_transform)

    def __worker_count_from_normalized(self, normalized_count: float) -> int:
        """
        Calculates how many workers are equivalent to the given normalized worker count (normalized to 10)

        :param normalized_count: the fraction normalized to 10 workers as company size
        :return: the actual count of workers
        """
        return int(math.ceil(len(self.__workerList) * normalized_count / 10))

    def __team_size_from_workers_and_team_fraction(self, normalized_count: float, team_size: float) -> float:
        """
        Calculates the size of the team it is allowed to have maximal

        :param normalized_count: the normalized worker count for the section
        :param team_size: the fraction of the team to the whole section
        :return: the head count of workers the team is allowed to have tops
        """
        return math.ceil(team_size * self.__worker_count_from_normalized(normalized_count))

    def __assign(self) -> None:
        """
        Assigns every worker in the stored worker list to a section in a stored section list
        """
        def qualifications_match(worker_skills: List[str], required_skills: List[str], allowed_missing: int = 0) -> bool:
            """
            Checks if the worker brings the skills that the section requires

            :param worker_skills: the qualifications / skills the worker has
            :param required_skills: the skills required by the section
            :param allowed_missing: a desperation factor of the section: the number of skills that can be missing
            :return: if the worker matches the sections requirements under the given condition (allowed_missing)
            """
            match_counter = 0
            for skill in required_skills:
                if skill in worker_skills:
                    match_counter += 1
            return match_counter + allowed_missing >= len(required_skills)

        def assign_to_sections(sections: Dict[Creator.DataStruct, int], worker_list: List[Creator.DataStruct],
                               allowed_missing: int) -> None:
            """
            Goes through the list of sections and assigns workers to them as long as they have a open position

            :param sections: the sections bundled with the count of open positions
            :param worker_list: the list of workers which look for a job
            :param allowed_missing: the desperation factor of the sections: says how many skills can be missing by the
                   worker to match the section
            """
            for sec in sections.keys():
                assigned_workers = []
                for worker in worker_list:
                    if sections[sec] == 0:
                        # the boat is full
                        break
                    if not qualifications_match(worker.skills, sec.skills, allowed_missing):
                        continue
                    # mark the worker as to be removed from the pool
                    assigned_workers.append(worker)
                    sections[sec] -= 1      # remove one vacant spot
                    self.__assignments.append((sec, worker))
                # remove all workers assigned
                for assigned in assigned_workers:
                    worker_list.remove(assigned)

        def count_open_positions(section_map: Dict[Creator.DataStruct, int]) -> int:
            """
            Goes through the count of open positions and sums them up

            :param section_map: the sections and their count of vacant positions
            :return: sum of all vacant positions across all sections
            """
            open_spots = 0
            for section_size in section_map.values():
                open_spots += section_size
            return open_spots

        # shuffle the workers to be simulate a more realistic result
        shuffled_workers = list(self.__workerList)
        random.shuffle(shuffled_workers)
        section_dict = {}
        for section in self.__sectionList:
            section_dict[section] = self.__worker_count_from_normalized(section.attributes["NormalizedWorkerCount"])
        # check if there's enough spots for every worker else crash to inform the user
        vacant = count_open_positions(section_dict)
        if vacant < len(shuffled_workers):
            raise AttributeError("Have {} vacant spaces but {} workers".format(vacant, len(shuffled_workers)))
        skill_mismatch_allowed = 0
        while len(shuffled_workers) > 0:    # stop when all workers are assigned -> some job might be staying open
            assign_to_sections(section_dict, shuffled_workers, skill_mismatch_allowed)
            skill_mismatch_allowed += 1

    @staticmethod
    def __val_to_string(val) -> str:
        """
        Converts the value to a string which usually defaults to the usage of pythons provided __str__() function except
        for lists
        :param val: the value to transform
        :return: a string representation of the value
        """
        if not isinstance(val, list):
            # just fallback to the python default
            return str(val)
        if not val:
            # means the list is empty
            return ""
        list_str = str(val[0])
        for i in range(1, len(val)):
            list_str += ", {}".format(val[i])
        return list_str


if __name__ == "__main__":
    c = Creator("../workers.json", "../sections.json")
    c.create_xlsx()
